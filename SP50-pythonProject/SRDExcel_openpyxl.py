import os

import Levenshtein
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import string
from scikitlearn import process_tfidf_similarity, process_indexIF
from Levenshtein import *
import copy


def moveZaoDian(strZaoDian):
    return strZaoDian.replace("有限公司", "") \
        .replace("Co., Ltd", "") \
        .replace("Co. Ltd", "") \
        .replace("Co.,Ltd", "") \
        .replace("Co.,Limited", "") \
        .replace("Co., Lt", "") \
        .replace("Co., Li", "")


def writeOneCell(curList, curLine, curCollom, cursheet, ShowAnyAll):
    curList = sorted(curList, key=(lambda x: x[0]), reverse=True)

    Enformula = ""
    intAll = 0
    curLineBaseLineNum = 3000
    smax = 10
    staartNum = smax * (curLine - 1) + curLineBaseLineNum + 1
    strFirstValue = ""
    for Enlist in curList:  # 第一个实例
        if (not (ShowAnyAll) and (intAll == 0) and (Enlist[0] >= 1)):
            return

        strAdd = Enlist[2] + "," + Enlist[1] + "," + str(Enlist[0])  # error
        # result = find_string(Enformula, strAdd)
        if (strAdd not in Enformula):
            intAll = intAll + 1
            Enformula = Enformula + "," + strAdd
            # Enformula = Enformula + "," + str(Enlist[0]) + "-------" + Enlist[2]
            cursheet.cell(row=staartNum + intAll, column=curCollom, value=strAdd)
            if (intAll == 1):
                strFirstValue = strAdd
            if (intAll > smax - 1):
                break;
    if (intAll == 0):
        return

    strExample = Enformula[1:]
    # strExample = "\'\"" + Enformula[1:] + "\"\'"
    # strExample = '=$A2,$A4' error
    # strExample = '=$A:$A' ok
    strOKExample = '"Dog,Cat,Bat,Bat"'  # ok

    #   chr(i)
    char2 = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u',
             'v', 'w', 'x', 'y', 'z']
    lettercur = char2[curCollom - 1]
    strOKExample = '=$' + lettercur + str(staartNum + 1) + ':$' + lettercur + str(staartNum + intAll)  # ok
    # cursheet.cell(row = curLine,column = curCollom,value = strExample)
    data_val = DataValidation(type="list",
                              formula1=strOKExample)  # You can change =$A:$A with a smaller range like =A1:A9
    cursheet.add_data_validation(data_val)
    strCell = lettercur + str(curLine)
    cursheet[strCell].value = strFirstValue
    data_val.add(
        cursheet[
            strCell])  # If you go to the cell B1 you will find a drop down list with all the values from the column A


def MainDo():
    wb = Workbook()
    ws = wb.create_sheet('New Sheet')
    for number in range(1, 100):  # Generates 99 "ip" address in the Column A;
        ws['A{}'.format(number)].value = "192.168.1.{}".format(number)
    data_val = DataValidation(type="list",
                              formula1='"Dog,Cat,Bat"')  # You can change =$A:$A with a smaller range like =A1:A9
    ws.add_data_validation(data_val)
    data_val.add(
        ws["B1"])  # If you go to the cell B1 you will find a drop down list with all the values from the column A
    wb.save('Test.xlsx')

    book = openpyxl.load_workbook(u'China_SRD_CP_Newgen_Mapping03.xlsx', data_only=True)
    sheetSRDCPCopy = book['SRD CP Copy']

    sheet = book['Local CP']

    bookWrite = openpyxl.load_workbook(u'China_SRD_CP_Newgen_Mapping03.xlsx')
    sheetWrite = bookWrite['Local CP']
    sheetLine = 0

    OneReadList = []
    rowint = 0
    for oneSRD_row in sheetSRDCPCopy.rows:
        if (rowint > 0):
            SRD = oneSRD_row[0].value
            SRD_COUNTERPARTY_ID = oneSRD_row[1].value
            OneReadList.append([SRD, SRD_COUNTERPARTY_ID])
        rowint = rowint + 1

    # 一行一行遍历
    for one_row in sheet.rows:
        ResultID = one_row[5]

        sheetLine = sheetLine + 1
        if (ResultID.value == None):
            #

            # //if (one_row.len>13):
            ResultID12Add = one_row[12]
            booladd12 = (ResultID12Add.value == None)
            # //if (one_row.len > 14):
            ResultID13Add = one_row[13]
            booladd13 = (ResultID13Add.value == None)

            if (booladd12 and booladd13):
                COUNTERPARTY_NAMEEN = one_row[9].value
                COUNTERPARTY_NAMECN = one_row[8].value

                Enlist = []  ## 空列表
                Cnlist = []  ## 空列表

                En1list = []  ## 空列表
                Cn1list = []  ## 空列表
                rowint = 0

                SaveFilename = 'China_00' + str(sheetLine) + ' COUNTERPARTY_NAMEEN(' + COUNTERPARTY_NAMEEN + ').xlsx'
                if not (os.path.exists(SaveFilename)):
                    print(
                        "maybeError COUNTERPARTY_NAMEEN=" + COUNTERPARTY_NAMEEN + "    COUNTERPARTY_NAMECN=" + COUNTERPARTY_NAMECN)

                    for oneSRD_OneRead in OneReadList:
                        SRD = oneSRD_OneRead[0]
                        SRD_COUNTERPARTY_ID = oneSRD_OneRead[1]
                        if not (SRD == None):
                            try:
                                COUNTERPARTY_NAME_EN_INT = process_tfidf_similarity(moveZaoDian(COUNTERPARTY_NAMEEN),
                                                                                    moveZaoDian(SRD))
                                COUNTERPARTY_NAME_CN_INT = process_tfidf_similarity(moveZaoDian(COUNTERPARTY_NAMECN),
                                                                                    moveZaoDian(SRD))

                                if (COUNTERPARTY_NAME_EN_INT > 0.5):
                                    Enlist.append([COUNTERPARTY_NAME_EN_INT, SRD, SRD_COUNTERPARTY_ID]);
                                if (COUNTERPARTY_NAME_CN_INT > 0.5):
                                    Cnlist.append([COUNTERPARTY_NAME_CN_INT, SRD, SRD_COUNTERPARTY_ID]);

                                En1 = Levenshtein.ratio(moveZaoDian(COUNTERPARTY_NAMEEN), moveZaoDian(SRD))
                                Cn1 = Levenshtein.ratio(moveZaoDian(COUNTERPARTY_NAMECN), moveZaoDian(SRD))
                                if (En1 > 0.5):
                                    En1list.append([En1, SRD, SRD_COUNTERPARTY_ID])
                                if (Cn1 > 0.5):
                                    Cn1list.append([Cn1, SRD, SRD_COUNTERPARTY_ID])
                            except:
                                print("error rowint=" + str(
                                    rowint) + ",COUNTERPARTY_NAMEEN=" + COUNTERPARTY_NAMEEN + ",SRD=" + SRD)

                        # print("sheetLine = " + str(sheetLine) + ",   rowint = " + str(rowint))
                        rowint = rowint + 1
                    print("bug " + str(sheetLine))
                    writeOneCell(Enlist, sheetLine, 7, sheetWrite, False)
                    writeOneCell(Cnlist, sheetLine, 8, sheetWrite, False)

                    writeOneCell(En1list, sheetLine, 13, sheetWrite, True)
                    writeOneCell(Cn1list, sheetLine, 14, sheetWrite, True)
                    bookWrite.save(SaveFilename)
        # break

    book.close()
    print("over")


def find_string(s, t):
    try:
        string.index(s, t)
        return True
    except(ValueError):
        return False


# pyinstaller -F SRDExcel_openpyxl.py
if __name__ == '__main__':
    MainDo()
